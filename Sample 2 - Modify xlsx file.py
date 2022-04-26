import os, sys
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import PatternFill



my_color = Font(color='ff0000')
my_fill = PatternFill(patternType='solid', fgColor='00ff00')

basePath = os.path.abspath(os.path.dirname(sys.argv[0]))
Source = os.path.join(basePath, 'Create an excel file.xlsx')

Data = load_workbook(Source, data_only=True)
# Loop for each sheet in the workbook
for Sheet in Data:
    #Modify the sheet name:
    print(Sheet.title)
    Sheet.title = 'New sheet name'
    # Loop for all available row
    for row in Sheet.iter_rows():
        # Loop for all cell in the row
        for cell in row:
            # Get the value of the cell
            CellValue = cell.value
            # Get the column letter of the cell
            CollLetter = cell.column_letter
            # Get the row number of the cell
            CollRow = cell.row
            # Get the address of the cell
            CellAddress = cell.coordinate 
            if CollRow % 2 == 0 and CellValue != None:
                print('CellAddress', CellAddress)
                cell.fill = my_fill
                cell.font  = my_color
# Overwrite file with edited file
Data.save(Source)